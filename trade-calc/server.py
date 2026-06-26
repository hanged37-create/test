import io
import os
import re
from datetime import datetime, date, timedelta

import requests
import openpyxl
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
import pytesseract
from PIL import Image
from flask import Flask, request, jsonify, send_from_directory, send_file, abort

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
VENDOR_DIR = os.path.normpath(os.path.join(BASE_DIR, "..", "vendor"))

app = Flask(__name__, static_folder=None)

# 로컬 winget 설치 기본 경로. PATH에 등록돼 있으면 그 값을 그대로 쓴다.
_DEFAULT_TESSERACT = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
if os.path.exists(os.environ.get("TESSERACT_CMD", _DEFAULT_TESSERACT)):
    pytesseract.pytesseract.tesseract_cmd = os.environ.get("TESSERACT_CMD", _DEFAULT_TESSERACT)

# SMBS(서울외국환중개) 통화 목록. (100) 표시가 있는 통화는 100단위 고시이므로
# 실제원화 = 주문외화 * 환율 / 100 으로 계산해야 한다.
CURRENCIES = [
    {"code": "USD", "name": "미국 달러", "per100": False},
    {"code": "JPY", "name": "일본 엔", "per100": True},
    {"code": "EUR", "name": "유로", "per100": False},
    {"code": "GBP", "name": "영국 파운드", "per100": False},
    {"code": "CAD", "name": "캐나다 달러", "per100": False},
    {"code": "CHF", "name": "스위스 프랑", "per100": False},
    {"code": "AUD", "name": "호주 달러", "per100": False},
    {"code": "NZD", "name": "뉴질랜드 달러", "per100": False},
    {"code": "CNY", "name": "중국 위안", "per100": False},
    {"code": "HKD", "name": "홍콩 달러", "per100": False},
    {"code": "TWD", "name": "대만 달러", "per100": False},
    {"code": "MNT", "name": "몽골 투그릭", "per100": False},
    {"code": "KZT", "name": "카자흐스탄 텡게", "per100": False},
    {"code": "THB", "name": "태국 밧", "per100": False},
    {"code": "SGD", "name": "싱가포르 달러", "per100": False},
    {"code": "IDR", "name": "인도네시아 루피아", "per100": True},
    {"code": "MYR", "name": "말레이시아 링깃", "per100": False},
    {"code": "PHP", "name": "필리핀 페소", "per100": False},
    {"code": "VND", "name": "베트남 동", "per100": True},
    {"code": "BND", "name": "브루나이 달러", "per100": False},
    {"code": "INR", "name": "인도 루피", "per100": False},
    {"code": "PKR", "name": "파키스탄 루피", "per100": False},
    {"code": "BDT", "name": "방글라데시 타카", "per100": False},
    {"code": "KHR", "name": "캄보디아 리엘", "per100": False},
    {"code": "MOP", "name": "마카오 파타카", "per100": False},
    {"code": "NPR", "name": "네팔 루피", "per100": False},
    {"code": "LKR", "name": "스리랑카 루피", "per100": False},
    {"code": "UZS", "name": "우즈베키스탄 숨", "per100": False},
    {"code": "MMK", "name": "미얀마 짯", "per100": False},
    {"code": "MXN", "name": "멕시코 페소", "per100": False},
    {"code": "BRL", "name": "브라질 헤알", "per100": False},
    {"code": "ARS", "name": "아르헨티나 페소", "per100": False},
    {"code": "CLP", "name": "칠레 페소", "per100": False},
    {"code": "COP", "name": "콜롬비아 페소", "per100": False},
    {"code": "SEK", "name": "스웨덴 크로나", "per100": False},
    {"code": "DKK", "name": "덴마크 크로네", "per100": False},
    {"code": "NOK", "name": "노르웨이 크로네", "per100": False},
    {"code": "RUB", "name": "러시아 루블", "per100": False},
    {"code": "HUF", "name": "헝가리 포린트", "per100": False},
    {"code": "PLN", "name": "폴란드 즈워티", "per100": False},
    {"code": "CZK", "name": "체코 코루나", "per100": False},
    {"code": "RON", "name": "루마니아 레우", "per100": False},
    {"code": "SAR", "name": "사우디아라비아 리얄", "per100": False},
    {"code": "QAR", "name": "카타르 리얄", "per100": False},
    {"code": "ILS", "name": "이스라엘 셰켈", "per100": False},
    {"code": "JOD", "name": "요르단 디나르", "per100": False},
    {"code": "KWD", "name": "쿠웨이트 디나르", "per100": False},
    {"code": "BHD", "name": "바레인 디나르", "per100": False},
    {"code": "AED", "name": "아랍에미리트 디르함", "per100": False},
    {"code": "TRY", "name": "튀르키예 리라", "per100": False},
    {"code": "OMR", "name": "오만 리얄", "per100": False},
    {"code": "ZAR", "name": "남아프리카공화국 랜드", "per100": False},
    {"code": "EGP", "name": "이집트 파운드", "per100": False},
    {"code": "KES", "name": "케냐 실링", "per100": False},
    {"code": "LYD", "name": "리비아 디나르", "per100": False},
    {"code": "ETB", "name": "에티오피아 비르", "per100": False},
    {"code": "FJD", "name": "피지 달러", "per100": False},
]
CURRENCY_BY_CODE = {c["code"]: c for c in CURRENCIES}

SMBS_XML_URL = "http://www.smbs.biz/ExRate/StdExRate_xml.jsp"
SET_RE = re.compile(r"label='(\d{2})\.(\d{2})\.(\d{2})'\s+value='([\d.]+)'")


def fetch_smbs_rates(currency, start_iso, end_iso):
    """SMBS 기간별 매매기준율 XML을 받아 [(YYYY-MM-DD, rate), ...] (날짜 오름차순)으로 반환."""
    arr_value = f"{currency}_{start_iso}_{end_iso}"
    resp = requests.get(
        SMBS_XML_URL,
        params={"arr_value": arr_value},
        timeout=10,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    resp.raise_for_status()
    text = resp.content.decode("euc-kr", errors="ignore")
    results = []
    for yy, mm, dd, val in SET_RE.findall(text):
        year = 2000 + int(yy)
        results.append((f"{year:04d}-{mm}-{dd}", float(val)))
    results.sort(key=lambda x: x[0])
    return results


def parse_date_cell(value, fallback_year=None):
    """엑셀 셀 값을 'YYYY-MM-DD' 문자열로 정규화. 연도가 없는 'MM-DD' 형식이면 fallback_year를 붙인다."""
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")

    s = str(value).strip().replace("/", "-").replace(".", "-").strip("-")
    for fmt in ("%Y-%m-%d", "%y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        parsed = datetime.strptime(s, "%m-%d")
        year = fallback_year or datetime.now().year
        return f"{int(year):04d}-{parsed.month:02d}-{parsed.day:02d}"
    except ValueError:
        return None


def find_header_row(ws):
    """'T.I 작성일' 헤더가 있는 행과 그 열 번호를 찾는다. '수리일' 컬럼이 같이 있으면 참고용으로 같이 잡아둔다
    (수리일은 이제 신고필증 OCR로 채우는 게 기본이라 엑셀에 없어도 된다)."""
    for row in ws.iter_rows(min_row=1, max_row=min(15, ws.max_row)):
        found = {}
        for cell in row:
            if cell.value is None:
                continue
            norm = str(cell.value).replace(" ", "").replace(".", "").replace("/", "")
            if "TI" in norm.upper() and "작성" in norm:
                found["ti"] = cell.column
            elif "수리" in norm:
                found["clearance"] = cell.column
        if "ti" in found:
            return row[0].row, found
    return None, {}


def parse_excel_rows(wb):
    ws = wb.active
    header_row, cols = find_header_row(ws)
    if header_row is None:
        raise ValueError("엑셀에서 'T.I 작성일' 컬럼을 찾을 수 없습니다. 헤더 표기를 확인해주세요.")

    rows = []
    seq = 1
    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        ti_cell = row[cols["ti"] - 1]
        cl_cell = row[cols["clearance"] - 1] if "clearance" in cols else None
        if ti_cell.value is None and (cl_cell is None or cl_cell.value is None):
            continue
        clearance_date = parse_date_cell(cl_cell.value) if cl_cell is not None else None
        ti_date = parse_date_cell(ti_cell.value, fallback_year=(clearance_date[:4] if clearance_date else None))
        rows.append({"seq": seq, "tiDate": ti_date, "clearanceDate": clearance_date})
        seq += 1
    return rows


def extract_text_from_pdf(file_bytes):
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    text_parts = []
    for page in doc:
        text = page.get_text().strip()
        if len(text) >= 20:
            text_parts.append(text)
        else:
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            text_parts.append(pytesseract.image_to_string(img, lang="kor+eng"))
    return "\n".join(text_parts)


def extract_text_from_image(file_bytes):
    img = Image.open(io.BytesIO(file_bytes))
    return pytesseract.image_to_string(img, lang="kor+eng")


_CURRENCY_PATTERN = re.compile(
    r"\b(" + "|".join(CURRENCY_BY_CODE.keys()) + r")[\s:\-]*([\d,]+\.?\d*)"
)
_DECLARATION_NO_PATTERN = re.compile(r"\b\d{5}-\d{2}-\d{6,7}[A-Z]?\b")
_CUSTOMS_DUTY_PATTERN = re.compile(r"관세\D{0,10}?([\d,]{3,})")
_VAT_PATTERN = re.compile(r"부가(?:가치)?세\D{0,10}?([\d,]{3,})")
_DATE_VALUE_PATTERN = r"(\d{4})[.\-/년]\s*(\d{1,2})[.\-/월]\s*(\d{1,2})"
_CLEARANCE_DATE_PATTERN = re.compile(r"수리일자?" + r"\D{0,10}?" + _DATE_VALUE_PATTERN)
_ISSUE_DATE_PATTERN = re.compile(r"작성일자?" + r"\D{0,10}?" + _DATE_VALUE_PATTERN)


def _extract_date(pattern, text):
    m = pattern.search(text)
    if not m:
        return None
    try:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return f"{y:04d}-{mo:02d}-{d:02d}"
    except ValueError:
        return None


def parse_declaration_fields(text):
    """수출/수입신고필증 텍스트에서 문서종류/신고번호/통화/금액/관세/수리일을 best-effort로 추출.
    관세는 신고필증에만 있고 세금계산서에는 없다."""
    result = {
        "documentType": None,
        "declarationNo": None,
        "currency": None,
        "foreignAmount": None,
        "customsDuty": None,
        "clearanceDate": None,
    }

    if "수출신고필증" in text:
        result["documentType"] = "export"
    elif "수입신고필증" in text:
        result["documentType"] = "import"

    m = _DECLARATION_NO_PATTERN.search(text)
    if m:
        result["declarationNo"] = m.group(0)

    m = _CURRENCY_PATTERN.search(text)
    if m:
        result["currency"] = m.group(1)
        try:
            result["foreignAmount"] = float(m.group(2).replace(",", ""))
        except ValueError:
            pass

    m = _CUSTOMS_DUTY_PATTERN.search(text)
    if m:
        try:
            result["customsDuty"] = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    result["clearanceDate"] = _extract_date(_CLEARANCE_DATE_PATTERN, text)

    result["rawTextPreview"] = text[:1500]
    return result


def parse_tax_invoice_fields(text):
    """수입세금계산서 텍스트에서 부가세/작성일을 best-effort로 추출. 관세는 여기 없다 (신고필증에만 있음)."""
    result = {"vatAmount": None, "issueDate": None}

    m = _VAT_PATTERN.search(text)
    if m:
        try:
            result["vatAmount"] = float(m.group(1).replace(",", ""))
        except ValueError:
            pass

    result["issueDate"] = _extract_date(_ISSUE_DATE_PATTERN, text)

    return result


@app.route("/api/currencies")
def api_currencies():
    return jsonify(CURRENCIES)


@app.route("/api/exchange-rate", methods=["POST"])
def api_exchange_rate():
    data = request.get_json(force=True) or {}
    currency = (data.get("currency") or "").upper().strip()
    date_str = (data.get("date") or "").strip()

    if currency not in CURRENCY_BY_CODE:
        return jsonify({"error": f"지원하지 않는 통화코드입니다: {currency}"}), 400
    try:
        target = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "date는 YYYY-MM-DD 형식이어야 합니다."}), 400

    start = target - timedelta(days=14)
    try:
        rates = fetch_smbs_rates(currency, start.isoformat(), target.isoformat())
    except requests.RequestException as e:
        return jsonify({"error": f"환율 조회 중 오류가 발생했습니다: {e}"}), 502

    target_str = target.isoformat()
    exact = next((r for r in rates if r[0] == target_str), None)
    if exact:
        matched_date, rate = exact
        is_fallback = False
    else:
        prior = [r for r in rates if r[0] < target_str]
        if not prior:
            return jsonify({"error": "해당 날짜의 매매기준율을 찾을 수 없습니다."}), 404
        matched_date, rate = prior[-1]
        is_fallback = True

    return jsonify(
        {
            "currency": currency,
            "rate": rate,
            "matchedDate": matched_date,
            "isFallback": is_fallback,
            "per100": CURRENCY_BY_CODE[currency]["per100"],
        }
    )


@app.route("/api/parse-excel", methods=["POST"])
def api_parse_excel():
    file = request.files.get("file")
    if not file:
        return jsonify({"error": "파일이 없습니다."}), 400
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
        rows = parse_excel_rows(wb)
    except Exception as e:
        return jsonify({"error": str(e)}), 400
    if not rows:
        return jsonify({"error": "엑셀에서 데이터 행을 찾을 수 없습니다."}), 400
    return jsonify({"rows": rows})


@app.route("/api/ocr-declaration", methods=["POST"])
def api_ocr_declaration():
    files = request.files.getlist("files") or request.files.getlist("file")
    if not files:
        return jsonify({"error": "파일이 없습니다."}), 400

    merged = {
        "documentType": None,
        "declarationNo": None,
        "currency": None,
        "foreignAmount": None,
        "customsDuty": None,
        "clearanceDate": None,
        "vatAmount": None,
        "taxInvoiceDate": None,
    }

    for file in files:
        filename = (file.filename or "").lower()
        content = file.read()
        try:
            if filename.endswith(".pdf"):
                text = extract_text_from_pdf(content)
            else:
                text = extract_text_from_image(content)
        except Exception as e:
            return jsonify({"error": f"문서 인식 중 오류가 발생했습니다 ({file.filename}): {e}"}), 500

        if "세금계산서" in text:
            tax_fields = parse_tax_invoice_fields(text)
            if tax_fields["vatAmount"] is not None:
                merged["vatAmount"] = tax_fields["vatAmount"]
            if tax_fields["issueDate"] is not None:
                merged["taxInvoiceDate"] = tax_fields["issueDate"]
        else:
            # 신고필증(수출/수입 공통). 관세·수리일은 신고필증에만 있다.
            decl_fields = parse_declaration_fields(text)
            for key in ("documentType", "declarationNo", "currency", "foreignAmount", "customsDuty", "clearanceDate"):
                if decl_fields.get(key) is not None:
                    merged[key] = decl_fields[key]

    return jsonify(merged)


@app.route("/api/export-excel", methods=["POST"])
def api_export_excel():
    data = request.get_json(force=True) or {}
    rows = data.get("rows", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "수출입계산"
    headers = ["구분", "T.I 작성일", "수리일", "환율", "주문외화", "실제원화", "관세", "원화+관세", "적요"]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r.get("label", ""),
                r.get("tiDate", ""),
                r.get("clearanceDate", ""),
                r.get("exchangeRate", ""),
                r.get("foreignAmount", ""),
                r.get("krwAmount", ""),
                r.get("customsDuty", ""),
                r.get("totalKrw", ""),
                r.get("note", ""),
            ]
        )
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"수출입계산_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


_JOURNAL_HEADERS = [
    "년도월일", "구분", "코드", "계정과목", "거래처코드", "거래처명", "적요코드", "적요명", "금액",
    "부서명", "사원명", "현장구분", "현장코드", "현장명", "거래처구분", "등록,계좌,카드번호",
    "업태", "종목", "우편번호", "주소", "상세주소", "대표자", "매입카드일때카드유형",
    "신용카드사코드", "카드사용여부", "프로젝트코드", "프로젝트명", "증빙불비 원인",
]


@app.route("/api/export-journal", methods=["POST"])
def api_export_journal():
    data = request.get_json(force=True) or {}
    rows = data.get("rows", [])

    import_rows = [r for r in rows if r.get("docType") == "import"]
    skipped_export_count = len(rows) - len(import_rows)
    if not import_rows:
        return jsonify({"error": "내보낼 수입 건이 없습니다. (수출 건은 일반전표 내보내기 대상이 아닙니다)"}), 400

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "일반전표"
    ws.append(_JOURNAL_HEADERS)
    pad = [""] * (len(_JOURNAL_HEADERS) - 9)

    for r in import_rows:
        ti_date = (r.get("tiDate") or "").replace("-", "")
        amount = r.get("totalKrw")
        amount = round(amount) if amount not in (None, "") else ""
        counterparty_code = r.get("counterpartyCode") or ""
        counterparty_name = r.get("counterpartyName") or ""
        note = r.get("note") or ""

        ws.append([ti_date, 1, 146, "상품", counterparty_code, counterparty_name, "", note, amount] + pad)
        ws.append([ti_date, 2, 251, "외상매입금", counterparty_code, counterparty_name, "", note, amount] + pad)

    for i in range(1, len(_JOURNAL_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"일반전표_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    resp = send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp.headers["X-Skipped-Export-Count"] = str(skipped_export_count)
    resp.headers["X-Written-Count"] = str(len(import_rows))
    return resp


@app.route("/")
def index():
    return send_from_directory(BASE_DIR, "index.html")


@app.route("/<path:path>")
def static_files(path):
    if path.startswith("vendor/"):
        return send_from_directory(VENDOR_DIR, path[len("vendor/"):])
    full = os.path.join(BASE_DIR, path)
    if os.path.isfile(full):
        return send_from_directory(BASE_DIR, path)
    abort(404)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8010, debug=False)
